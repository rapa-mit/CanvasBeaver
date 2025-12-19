#!/usr/bin/env python3
"""
grade_processor.py
--------------------------------
Advanced grade processing for Canvas courses using the Canvas API.

This module processes Canvas gradebook data and computes final grades
based on configurable grade type weights. It replaces CSV-based processing
with direct API queries.

Features:
- Configurable grade type weights (flexible partial semester grading)
- Automatic drop lowest grades by type
- Letter grade computation with configurable scale
- Anomaly detection for suspicious grade patterns
- Individual student reports
- Summary reports and analytics
"""
from __future__ import annotations

import statistics
import yaml
from bisect import bisect
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from .gradebook import Gradebook, StudentSummary

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class BadGradeTypesWeight(Exception):
    """Raised when grade type weights don't sum to expected value."""
    pass


class GradeConfiguration:
    """Manages grade type configuration and validation."""

    def __init__(self, grade_types: Dict[str, float], allow_partial: bool = False):
        """Initialize grade configuration.

        Args:
            grade_types: Dictionary mapping assignment group names to weights (0.0-1.0)
            allow_partial: If True, allows weights to sum to less than 1.0 for mid-semester
        """
        self.grade_types = grade_types
        self.allow_partial = allow_partial
        self._total_weight = sum(grade_types.values())

        # Validate weights
        if self._total_weight > 1.0 + 1e-8:
            raise BadGradeTypesWeight(
                f"Grade type weights sum to {self._total_weight:.3f}, cannot exceed 1.0"
            )
        elif self._total_weight < 1e-8:
            raise BadGradeTypesWeight("Grade type weights must sum to greater than 0")

        if allow_partial:
            if abs(self._total_weight - 1.0) > 1e-8:
                print(
                    f"Note: Partial semester grading - weights sum to {self._total_weight:.1%}"
                )
        else:
            if abs(self._total_weight - 1.0) > 1e-8:
                raise BadGradeTypesWeight(
                    f"Grade type weights sum to {self._total_weight:.3f}, expected 1.0"
                )

    @property
    def total_weight(self) -> float:
        """Return the total weight of all grade types."""
        return self._total_weight

    @property
    def is_partial(self) -> bool:
        """Return True if this is partial semester grading."""
        return self.allow_partial and abs(self._total_weight - 1.0) > 1e-8

    def weight_for_type(self, type_name: str) -> float:
        """Get weight for a grade type."""
        return self.grade_types.get(type_name, 0.0)


def load_letter_grade_scale(scale_file: Union[str, Path]) -> Dict[float, str]:
    """Load letter grade scale from YAML file.
    
    Args:
        scale_file: Path to YAML file with grade scale
        
    Returns:
        Dictionary mapping thresholds to letter grades
    """
    with open(scale_file, 'r') as f:
        data = yaml.safe_load(f)
    return data['scale']


def letter_grade(percentage: float, scale: Optional[Dict[float, str]] = None) -> str:
    """Convert a numerical grade (percentage) to a letter grade.

    Args:
        percentage: Grade as a percentage (0.0 to 1.0)
        scale: Optional custom grading scale as {threshold: grade}
               If None, loads from MIT-letter-grades.yaml

    Returns:
        Letter grade string (e.g., 'A', 'B+', 'C-', etc.)
    """
    if scale is None:
        # Load default MIT scale
        scale_path = Path(__file__).parent.parent / 'MIT-letter-grades.yaml'
        if scale_path.exists():
            scale = load_letter_grade_scale(scale_path)
        else:
            # Fallback to hardcoded scale if file not found
            scale = {
                0.00: 'F',
                0.61: 'D-',
                0.70: 'D',
                0.74: 'C-',
                0.77: 'C',
                0.80: 'C+',
                0.84: 'B-',
                0.87: 'B',
                0.90: 'B+',
                0.94: 'A-',
                0.97: 'A',
                1.00: 'A+',
            }
    
    # Sort scale by threshold
    sorted_items = sorted(scale.items())
    breakpoints = [t for t, _ in sorted_items]
    letters = [g for _, g in sorted_items]

    # Find appropriate grade using bisect
    # bisect(breakpoints, percentage) returns the insertion point
    # We want the grade corresponding to the highest threshold <= percentage
    i = bisect(breakpoints, percentage)
    
    # Clamp index to valid range
    # i can be 0 (below all thresholds) to len(breakpoints) (above all thresholds)
    # We want index 0 to len(letters)-1
    index = max(0, min(i - 1, len(letters) - 1))
    
    # Special case: if percentage is below the lowest threshold, still use lowest grade
    if percentage < breakpoints[0]:
        index = 0
    
    return letters[index]


@dataclass
class GradeTypeResult:
    """Results for one grade type."""
    type_name: str
    assignments: Dict[str, Tuple[float, float]]  # name -> (percentage, points)
    average_percentage: float
    contribution_to_total: float
    num_assignments: int
    dropped: Optional[Tuple[str, float, float]] = None  # (name, percentage, points)


@dataclass
class ProcessedStudent:
    """Student with processed grades and computed final grade."""
    user_id: Any
    name: str
    email: Optional[str] = None
    mit_id: Optional[str] = None
    
    grade_types: Dict[str, GradeTypeResult] = field(default_factory=dict)
    final_percentage: float = 0.0
    normalized_percentage: float = 0.0
    letter_grade: str = 'F'
    modified_letter_grade: Optional[str] = None  # Alternative letter grade with modified scale
    anomalies: List[str] = field(default_factory=list)
    weight_normalization_factor: float = 1.0  # Track normalization for reporting

    def format_report(self, show_anomalies: bool = True, partial_semester: bool = False) -> str:
        """Generate formatted individual grade report."""
        line = "=" * 70
        s = f"{line}\n"
        s += "GRADE REPORT"
        if partial_semester:
            s += " (PARTIAL SEMESTER - Based on graded assignments only)"
        s += "\n"
        s += f"{line}\n\n"
        s += f"Student: {self.name}\n"
        if self.email:
            s += f"Email:   {self.email}\n"
        if self.mit_id:
            s += f"MIT ID:  {self.mit_id}\n"
        s += f"\n{line}\n"
        
        if partial_semester:
            # Show current grade based on graded work
            s += f"CURRENT GRADE (on graded work): {self.normalized_percentage*100:.2f}%\n"
            s += f"LETTER GRADE ACCORDING TO COURSE GRADING SCHEME: {self.letter_grade}\n"
            if self.modified_letter_grade:
                s += f"LETTER GRADE SUBMITTED TO REGISTRAR ACCORDING TO MODIFIED CUTOFFS: {self.modified_letter_grade}\n"
            s += f"(Graded categories worth {self.weight_normalization_factor*100:.0f}% of course, normalized to 100%)\n"
        else:
            s += f"OVERALL GRADE: {self.normalized_percentage*100:.2f}%\n"
            s += f"LETTER GRADE ACCORDING TO COURSE GRADING SCHEME: {self.letter_grade}\n"
            if self.modified_letter_grade:
                s += f"LETTER GRADE SUBMITTED TO REGISTRAR ACCORDING TO MODIFIED CUTOFFS: {self.modified_letter_grade}\n"
        
        s += f"{line}\n\n"

        # Grade breakdown by type
        for type_name, result in self.grade_types.items():
            s += f"{type_name.upper()}\n"
            s += f"{'-' * len(type_name)}\n"

            # Individual assignments
            for assgn_name, (pct, pts) in result.assignments.items():
                clean_name = assgn_name.split('(')[0].strip()
                s += f"  {clean_name:45s} {pct*100:6.2f}%  ({pts:.2f} pts)\n"

            # Summary for this type
            s += f"\n  Average across {result.num_assignments} assignment(s): "
            s += f"{result.average_percentage*100:.2f}%\n"
            
            # Show the weighted contribution (performance × category weight)
            s += f"  Weighted contribution: {result.contribution_to_total*100:.2f}%\n"
            
            if partial_semester:
                # Also show what this represents as % of current graded work
                normalized_contribution = result.contribution_to_total / self.weight_normalization_factor
                s += f"  As % of graded work: {normalized_contribution*100:.2f}%\n"

            # Show dropped grade if any
            if result.dropped:
                dropped_name, dropped_pct, dropped_pts = result.dropped
                clean_name = dropped_name.split('(')[0].strip()
                s += f"  ** Lowest grade dropped: {clean_name} "
                s += f"({dropped_pct*100:.2f}%, {dropped_pts:.2f} pts)\n"

            s += "\n"

        # Display anomalies if any
        if show_anomalies and self.anomalies:
            s += f"\n{'!' * 70}\n"
            s += "GRADE PATTERN ALERTS\n"
            s += f"{'!' * 70}\n"
            for anomaly in self.anomalies:
                s += f"  ⚠ {anomaly}\n"
            s += f"{'!' * 70}\n"

        return s


class GradeProcessor:
    """Process grades from Canvas gradebook with configurable weights."""

    def __init__(
        self,
        gradebook: Gradebook,
        grade_config: Optional[GradeConfiguration] = None,
        assignment_groups: Optional[List[Any]] = None,
        letter_grade_scale: Optional[Union[str, Path, Dict[float, str]]] = None,
        modified_grade_scale: Optional[Union[str, Path, Dict[float, str]]] = None,
    ):
        """Initialize grade processor.

        Args:
            gradebook: Loaded Gradebook object
            grade_config: Optional grade configuration with weights (if not using Canvas groups)
            assignment_groups: Optional list of Canvas assignment groups with weights
            letter_grade_scale: Optional path to letter grade YAML file or dict of scale
            modified_grade_scale: Optional path to modified letter grade YAML file or dict of scale
            
        Note: Either grade_config OR assignment_groups should be provided.
              If assignment_groups is provided, weights are read from Canvas.
        """
        self.gradebook = gradebook
        self.processed_students: Dict[Any, ProcessedStudent] = {}
        
        # Load letter grade scale
        if letter_grade_scale is None:
            self.letter_grade_scale = None  # Will use default
        elif isinstance(letter_grade_scale, dict):
            self.letter_grade_scale = letter_grade_scale
        else:
            self.letter_grade_scale = load_letter_grade_scale(letter_grade_scale)
        
        # Load modified letter grade scale
        if modified_grade_scale is None:
            self.modified_grade_scale = None
        elif isinstance(modified_grade_scale, dict):
            self.modified_grade_scale = modified_grade_scale
        else:
            self.modified_grade_scale = load_letter_grade_scale(modified_grade_scale)

        # Build configuration from Canvas assignment groups or use provided config
        if assignment_groups is not None:
            self._init_from_canvas_groups(assignment_groups)
        elif grade_config is not None:
            self.config = grade_config
            self._assignment_to_type = self._build_assignment_map_legacy()
        else:
            raise ValueError("Either grade_config or assignment_groups must be provided")

    def _init_from_canvas_groups(self, assignment_groups: List[Any]) -> None:
        """Initialize configuration from Canvas assignment groups.
        
        Args:
            assignment_groups: List of Canvas assignment group objects with weights
        """
        # Store assignment groups for later reference (e.g., Excel export)
        self._assignment_groups = assignment_groups
        
        # Extract weights from Canvas groups (Canvas uses percentages, we need decimals)
        grade_types = {}
        self._group_id_to_name: Dict[int, str] = {}
        
        for group in assignment_groups:
            name = getattr(group, 'name', None)
            group_id = getattr(group, 'id', None)
            weight = getattr(group, 'group_weight', None)
            
            if name and weight is not None:
                # Canvas stores weights as percentages (e.g., 40.0 for 40%)
                # Convert to decimal (0.40)
                grade_types[name] = weight / 100.0
                self._group_id_to_name[group_id] = name
        
        # Create configuration (allow partial since Canvas might not have all groups weighted)
        self.config = GradeConfiguration(grade_types, allow_partial=True)
        
        # Build assignment mapping using assignment_group_id
        self._assignment_to_type: Dict[int, str] = {}
        for aid, assignment in self.gradebook.assignments.items():
            group_id = assignment.assignment_group_id
            if group_id in self._group_id_to_name:
                self._assignment_to_type[aid] = self._group_id_to_name[group_id]
    
    def _identify_graded_assignments(self) -> set[int]:
        """Identify assignments that have been graded (at least one non-zero score).
        
        Returns:
            Set of assignment IDs that have been graded
        """
        graded = set()
        for aid in self.gradebook.assignments.keys():
            # Check if any student has a non-zero score for this assignment
            for student in self.gradebook.get_students():
                score_obj = student.scores.get(aid)
                if score_obj and not score_obj.excused:
                    if score_obj.score is not None and score_obj.score > 0:
                        graded.add(aid)
                        break
        return graded
    
    def _calculate_graded_category_weights(self) -> None:
        """Calculate which categories have graded work and their weights.
        
        Simple logic: If a category has ANY graded assignments, use its full weight.
        """
        # Identify which categories have graded assignments
        graded_categories = set()
        for aid in self._graded_assignments:
            type_name = self._assignment_to_type.get(aid)
            if type_name:
                graded_categories.add(type_name)
        
        # Sum the weights of categories that have graded work
        self._active_weight_total = 0.0
        
        print("\nCategories with graded assignments:")
        for type_name in sorted(self.config.grade_types.keys()):
            weight = self.config.grade_types[type_name]
            if type_name in graded_categories:
                self._active_weight_total += weight
                print(f"  {type_name}: weight = {weight*100:.1f}%")
            else:
                print(f"  {type_name}: weight = {weight*100:.1f}% (NO GRADED WORK - excluded)")
        
        print(f"\nTotal weight of graded categories: {self._active_weight_total*100:.1f}%")
        print(f"Will normalize to 100%")
        
        # Store the normalization factor
        self._weight_normalization = self._active_weight_total if self._active_weight_total > 0 else 1.0

    def _build_assignment_map_legacy(self) -> Dict[int, str]:
        """Build assignment to type mapping using pattern matching (legacy mode).
        
        This is used when grade_config is provided instead of Canvas assignment groups.
        Requires group_name_map to be set in config.
        """
        mapping: Dict[int, str] = {}
        # This method needs group_name_map which should be in config
        # For now, return empty - will need to refactor config to include this
        return mapping

    def process_grades(
        self,
        drop_lowest: Optional[Dict[str, int]] = None,
        detect_anomalies: bool = True,
        only_graded: bool = True,
    ) -> None:
        """Process all student grades.

        Args:
            drop_lowest: Dict mapping grade type names to number of assignments to drop
            detect_anomalies: Whether to run anomaly detection
            only_graded: If True, only include assignments that have been graded (partial semester)
        """
        drop_lowest = drop_lowest or {}
        
        # Identify which assignments have been graded
        if only_graded:
            self._graded_assignments = self._identify_graded_assignments()
            print(f"\nPartial semester grading: {len(self._graded_assignments)} of {len(self.gradebook.assignments)} assignments have been graded")
            
            # Calculate weights for categories with graded work
            self._calculate_graded_category_weights()
        else:
            self._graded_assignments = set(self.gradebook.assignments.keys())

        # Process each student
        for student in self.gradebook.get_students():
            # Skip test students
            if student.name and ('Perfect' in student.name or 'Test Student' in student.name):
                continue

            processed = self._process_student(student, drop_lowest)
            self.processed_students[student.id] = processed

        # Anomaly detection
        if detect_anomalies:
            self._detect_anomalies()

    def _process_student(
        self,
        student: StudentSummary,
        drop_lowest: Dict[str, int],
    ) -> ProcessedStudent:
        """Process grades for a single student."""
        processed = ProcessedStudent(
            user_id=student.id,
            name=student.name or str(student.id),
            email=student.email,
            mit_id=student.login,  # Assuming login is MIT ID
        )

        # Group assignments by type (only graded assignments)
        assignments_by_type: Dict[str, Dict[int, Tuple[float, float]]] = {}
        for aid, score in student.scores.items():
            # Skip if not in graded assignments
            if not hasattr(self, '_graded_assignments') or aid not in self._graded_assignments:
                continue
                
            if score.excused:
                continue
            
            type_name = self._assignment_to_type.get(aid)
            if not type_name:
                continue

            if type_name not in assignments_by_type:
                assignments_by_type[type_name] = {}

            # Get percentage and points
            if score.score is not None and score.points_possible and score.points_possible > 0:
                pct = score.score / score.points_possible
                pts = score.score
            else:
                pct = 0.0
                pts = 0.0

            assignments_by_type[type_name][aid] = (pct, pts)

        # Process each grade type
        for type_name, assignments in assignments_by_type.items():
            # Sort by score (for dropping lowest)
            sorted_items = sorted(assignments.items(), key=lambda x: x[1][0])
            
            # Drop lowest if requested
            num_to_drop = drop_lowest.get(type_name, 0)
            dropped = None
            if num_to_drop > 0 and len(sorted_items) > num_to_drop:
                # Keep dropped info
                dropped_aid, (dropped_pct, dropped_pts) = sorted_items[0]
                dropped_name = self.gradebook.assignments[dropped_aid].name
                dropped = (dropped_name, dropped_pct, dropped_pts)
                
                # Remove lowest
                sorted_items = sorted_items[num_to_drop:]

            # Compute average and contribution
            if sorted_items:
                total_pct = sum(pct for _, (pct, _) in sorted_items)
                avg_pct = total_pct / len(sorted_items)
                
                # Use the full category weight (no proportional scaling)
                weight = self.config.weight_for_type(type_name)
                    
                contribution = avg_pct * weight

                # Build assignment dict with names
                assgn_dict = {}
                for aid, (pct, pts) in sorted_items:
                    name = self.gradebook.assignments[aid].name
                    assgn_dict[name] = (pct, pts)

                result = GradeTypeResult(
                    type_name=type_name,
                    assignments=assgn_dict,
                    average_percentage=avg_pct,
                    contribution_to_total=contribution,
                    num_assignments=len(sorted_items),
                    dropped=dropped,
                )
                processed.grade_types[type_name] = result
                processed.final_percentage += contribution

        # Normalize and compute letter grade
        # For partial semester, normalize by the total partial weight
        if hasattr(self, '_weight_normalization') and self._weight_normalization > 0:
            processed.normalized_percentage = processed.final_percentage / self._weight_normalization
            processed.weight_normalization_factor = self._weight_normalization
        elif self.config.total_weight > 0:
            processed.normalized_percentage = processed.final_percentage / self.config.total_weight
            processed.weight_normalization_factor = self.config.total_weight
        else:
            processed.normalized_percentage = 0.0
            processed.weight_normalization_factor = 1.0

        processed.letter_grade = letter_grade(
            processed.normalized_percentage,
            scale=self.letter_grade_scale
        )
        
        # Compute modified letter grade if modified scale is provided
        if self.modified_grade_scale:
            processed.modified_letter_grade = letter_grade(
                processed.normalized_percentage,
                scale=self.modified_grade_scale
            )

        return processed

    def _detect_anomalies(self) -> None:
        """Detect anomalous grade patterns.

        Flags:
        1. Large gaps between grade types (high in one, low in another)
        2. High variance within a grade type
        3. Statistical outliers compared to class
        """
        # Calculate class statistics by grade type
        type_stats: Dict[str, Dict[str, float]] = {}
        for type_name in self.config.grade_types.keys():
            percentages = []
            for student in self.processed_students.values():
                if type_name in student.grade_types:
                    percentages.append(student.grade_types[type_name].average_percentage)

            if len(percentages) >= 2:
                type_stats[type_name] = {
                    'mean': statistics.mean(percentages),
                    'stdev': statistics.stdev(percentages),
                }

        # Check each student
        for student in self.processed_students.values():
            # Check 1: Large gaps between types
            type_percentages = {
                name: result.average_percentage
                for name, result in student.grade_types.items()
            }

            if len(type_percentages) >= 2:
                for type1, pct1 in type_percentages.items():
                    for type2, pct2 in type_percentages.items():
                        if type1 >= type2:
                            continue
                        gap = abs(pct1 - pct2)
                        if max(pct1, pct2) > 0.90 and gap > 0.20:
                            high_type = type1 if pct1 > pct2 else type2
                            low_type = type2 if pct1 > pct2 else type1
                            high_pct = max(pct1, pct2)
                            low_pct = min(pct1, pct2)
                            student.anomalies.append(
                                f"{high_type} avg is {high_pct*100:.1f}% but "
                                f"{low_type} avg is only {low_pct*100:.1f}% (gap: {gap*100:.1f}%)"
                            )

            # Check 2: High variance within a type
            for type_name, result in student.grade_types.items():
                if result.num_assignments >= 3:
                    percentages = [pct for pct, _ in result.assignments.values()]
                    variance = statistics.stdev(percentages)
                    mean = statistics.mean(percentages)

                    if variance > 0.20 and mean > 0.30:
                        student.anomalies.append(
                            f"High variance in {type_name} scores "
                            f"(stdev: {variance*100:.1f}%, mean: {mean*100:.1f}%)"
                        )

            # Check 3: Statistical outliers
            for type_name, result in student.grade_types.items():
                if type_name in type_stats and type_stats[type_name]['stdev'] > 0:
                    stats = type_stats[type_name]
                    z_score = (result.average_percentage - stats['mean']) / stats['stdev']
                    if z_score > 2.0 and result.average_percentage > 0.95:
                        student.anomalies.append(
                            f"Statistical outlier in {type_name}: "
                            f"{result.average_percentage*100:.1f}% "
                            f"(class mean: {stats['mean']*100:.1f}%, z-score: {z_score:.2f})"
                        )

    def get_sorted_students(
        self,
        by: str = 'name',
        ascending: bool = True,
    ) -> List[ProcessedStudent]:
        """Get sorted list of processed students.

        Args:
            by: Sort key ('name', 'grade', 'percentage')
            ascending: Sort order
        """
        students = list(self.processed_students.values())

        if by == 'name':
            students.sort(key=lambda s: s.name, reverse=not ascending)
        elif by == 'grade':
            students.sort(
                key=lambda s: s.normalized_percentage,
                reverse=not ascending,
            )
        elif by == 'percentage':
            students.sort(
                key=lambda s: s.normalized_percentage,
                reverse=not ascending,
            )

        return students

    def print_summary(self) -> None:
        """Print summary statistics."""
        students = list(self.processed_students.values())
        if not students:
            print("No students processed")
            return

        percentages = [s.normalized_percentage for s in students]
        print("\n" + "=" * 70)
        print("GRADE SUMMARY")
        print("=" * 70)
        print(f"Total students: {len(students)}")
        print(f"Mean: {statistics.mean(percentages)*100:.2f}%")
        print(f"Median: {statistics.median(percentages)*100:.2f}%")
        if len(percentages) > 1:
            print(f"Std Dev: {statistics.stdev(percentages)*100:.2f}%")
        print(f"Min: {min(percentages)*100:.2f}%")
        print(f"Max: {max(percentages)*100:.2f}%")

        # Letter grade distribution
        letter_counts: Dict[str, int] = {}
        for s in students:
            letter_counts[s.letter_grade] = letter_counts.get(s.letter_grade, 0) + 1

        print("\nLetter Grade Distribution:")
        for letter in ['A+', 'A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D', 'D-', 'F']:
            count = letter_counts.get(letter, 0)
            if count > 0:
                print(f"  {letter}: {count}")

        # Anomalies
        flagged = sum(1 for s in students if s.anomalies)
        if flagged > 0:
            print(f"\nStudents with grade pattern alerts: {flagged}")

        print("=" * 70)

    def save_individual_reports(self, directory: str = 'individual-grades', include_excel: bool = True) -> None:
        """Save individual student reports to files.
        
        Args:
            directory: Directory to save reports
            include_excel: If True, also create Excel files with formulas
        """
        dir_path = Path(directory)
        dir_path.mkdir(exist_ok=True)

        # Clean existing files
        for f in dir_path.glob('*.txt'):
            f.unlink()
        if include_excel:
            for f in dir_path.glob('*.xlsx'):
                f.unlink()

        # Check if this is partial semester grading
        partial = hasattr(self, '_graded_assignments')

        for student in self.processed_students.values():
            # Save text report
            filename = dir_path / f"{student.name}.txt"
            filename.write_text(student.format_report(partial_semester=partial), encoding='utf-8')
            
            # Save Excel report if requested and library available
            if include_excel and HAS_OPENPYXL:
                excel_filename = dir_path / f"{student.name}.xlsx"
                self._save_student_excel(student, excel_filename, partial)

        report_types = "text and Excel" if (include_excel and HAS_OPENPYXL) else "text"
        print(f"Saved {len(self.processed_students)} individual {report_types} reports to {directory}/")

    def _save_student_excel(self, student: ProcessedStudent, filename: Path, partial_semester: bool) -> None:
        """Create an Excel file with student grades and formulas.
        
        Args:
            student: ProcessedStudent object
            filename: Path to save Excel file
            partial_semester: Whether this is partial semester grading
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Grade Report"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "GRADE REPORT"
        if partial_semester:
            cell.value += " (PARTIAL SEMESTER)"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Student info
        ws[f'A{row}'] = "Student:"
        ws[f'B{row}'] = student.name
        ws[f'A{row}'].font = bold_font
        row += 1
        
        if student.email:
            ws[f'A{row}'] = "Email:"
            ws[f'B{row}'] = student.email
            ws[f'A{row}'].font = bold_font
            row += 1
        
        if student.mit_id:
            ws[f'A{row}'] = "MIT ID:"
            ws[f'B{row}'] = student.mit_id
            ws[f'A{row}'].font = bold_font
            row += 1
        
        row += 1
        
        # Overall grade summary header row
        if partial_semester:
            ws[f'A{row}'] = "CURRENT GRADE (on graded work):"
        else:
            ws[f'A{row}'] = "FINAL GRADE:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = subheader_fill
        
        # This will be filled with formula later after we know the contribution rows
        ws[f'D{row}'].font = Font(bold=True, size=12)
        ws[f'D{row}'].fill = subheader_fill
        ws[f'D{row}'].number_format = '0.00"%"'
        
        ws[f'E{row}'].font = Font(bold=True, size=12)
        ws[f'E{row}'].fill = subheader_fill
        
        grade_summary_row = row
        row += 2
        
        # Track row numbers for formula references
        category_contribution_rows = {}
        
        # For each grade type
        for type_name, result in student.grade_types.items():
            # Category header
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = type_name.upper()
            cell.font = subheader_font
            cell.fill = subheader_fill
            row += 1
            
            # Assignment table headers
            ws[f'A{row}'] = "Assignment"
            ws[f'B{row}'] = "Score"
            ws[f'C{row}'] = "Max Points"
            ws[f'D{row}'] = "Percentage"
            ws[f'E{row}'] = "Included"
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1
            
            first_assignment_row = row
            
            # Assignments
            for assignment_name, (percentage, points) in result.assignments.items():
                # Look up max points from gradebook
                max_points = None
                for aid, assignment in self.gradebook.assignments.items():
                    if assignment.name == assignment_name:
                        max_points = assignment.points_possible
                        break
                
                # Fallback if not found: calculate from percentage (percentage is 0-1, not 0-100)
                if max_points is None or max_points == 0:
                    max_points = points / percentage if percentage > 0 else 0
                
                # Check if this is the dropped assignment
                is_dropped = (result.dropped and result.dropped[0] == assignment_name)
                
                ws[f'A{row}'] = assignment_name
                ws[f'B{row}'] = points
                ws[f'C{row}'] = max_points
                # Formula for percentage
                ws[f'D{row}'] = f"=IF(C{row}>0, B{row}/C{row}*100, 0)"
                ws[f'D{row}'].number_format = '0.00"%"'
                ws[f'E{row}'] = "No (Dropped)" if is_dropped else "Yes"
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].border = thin_border
                
                row += 1
            
            # Check if dropped assignment exists
            if result.dropped:
                dropped_name, dropped_pct, dropped_pts = result.dropped
                # Check if it's already in assignments (shouldn't be)
                if dropped_name not in result.assignments:
                    # Look up max points from gradebook
                    max_points = None
                    for aid, assignment in self.gradebook.assignments.items():
                        if assignment.name == dropped_name:
                            max_points = assignment.points_possible
                            break
                    
                    # Fallback: calculate from percentage (percentage is 0-1)
                    if max_points is None or max_points == 0:
                        max_points = dropped_pts / dropped_pct if dropped_pct > 0 else 0
                    
                    ws[f'A{row}'] = dropped_name
                    ws[f'B{row}'] = dropped_pts
                    ws[f'C{row}'] = max_points
                    ws[f'D{row}'] = f"=IF(C{row}>0, B{row}/C{row}*100, 0)"
                    ws[f'D{row}'].number_format = '0.00"%"'
                    ws[f'E{row}'] = "No (Dropped)"
                    
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].border = thin_border
                    
                    row += 1
            
            last_assignment_row = row - 1
            
            # Average calculation
            ws[f'A{row}'] = "Average:"
            ws[f'A{row}'].font = bold_font
            
            # Formula: AVERAGEIF to only include rows where column E = "Yes"
            ws[f'D{row}'] = f'=AVERAGEIF(E{first_assignment_row}:E{last_assignment_row},"Yes",D{first_assignment_row}:D{last_assignment_row})'
            ws[f'D{row}'].number_format = '0.00"%"'
            ws[f'D{row}'].font = bold_font
            ws[f'D{row}'].border = thin_border
            
            average_row = row
            row += 1
            
            # Category weight and contribution
            category_weight = self._get_category_weight(type_name)
            ws[f'A{row}'] = f"Category Weight:"
            ws[f'A{row}'].font = bold_font
            ws[f'D{row}'] = category_weight
            ws[f'D{row}'].number_format = '0.00"%"'
            ws[f'D{row}'].border = thin_border
            weight_row = row
            row += 1
            
            ws[f'A{row}'] = f"Contribution to Total:"
            ws[f'A{row}'].font = bold_font
            # Formula: average * weight / 100
            ws[f'D{row}'] = f'=D{average_row}*D{weight_row}/10000'
            ws[f'D{row}'].number_format = '0.00"%"'
            ws[f'D{row}'].font = bold_font
            ws[f'D{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws[f'D{row}'].border = thin_border
            
            category_contribution_rows[type_name] = row
            row += 2
        
        # Raw total calculation (sum of contributions)
        ws[f'A{row}'] = "TOTAL (weighted sum):"
        ws[f'A{row}'].font = bold_font
        
        # Sum all category contributions
        contribution_cells = [f"D{r}" for r in category_contribution_rows.values()]
        ws[f'D{row}'] = f'={"+".join(contribution_cells)}'
        ws[f'D{row}'].number_format = '0.00"%"'
        ws[f'D{row}'].font = bold_font
        ws[f'D{row}'].border = thin_border
        
        raw_total_row = row
        row += 1
        
        # Normalization factor for partial semester
        if partial_semester:
            ws[f'A{row}'] = "Normalization factor:"
            ws[f'A{row}'].font = bold_font
            ws[f'D{row}'] = student.weight_normalization_factor
            ws[f'D{row}'].number_format = '0.0000'
            ws[f'D{row}'].border = thin_border
            normalization_row = row
            row += 1
        
        # Final normalized grade
        ws[f'A{row}'] = "CURRENT GRADE:" if partial_semester else "FINAL GRADE:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Normalized grade formula
        if partial_semester:
            ws[f'D{row}'] = f'=D{raw_total_row}/D{normalization_row}'
        else:
            ws[f'D{row}'] = f'=D{raw_total_row}'
        ws[f'D{row}'].number_format = '0.00"%"'
        ws[f'D{row}'].font = Font(bold=True, size=12)
        ws[f'D{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws[f'D{row}'].border = thin_border
        
        current_grade_row = row
        
        # Now fill in the grade summary row with formula reference to current grade
        ws[f'D{grade_summary_row}'] = f'=D{current_grade_row}'
        ws[f'D{grade_summary_row}'].number_format = '0.00"%"'
        
        # Add letter grade with VLOOKUP to conversion table
        ws[f'E{grade_summary_row}'] = f'=VLOOKUP(D{grade_summary_row},Grades!A:B,2,TRUE)'
        
        # Add modified letter grade if modified scale is provided
        if self.modified_grade_scale:
            ws[f'F{grade_summary_row}'] = "Modified Grade:"
            ws[f'F{grade_summary_row}'].font = Font(bold=True, size=12)
            ws[f'F{grade_summary_row}'].fill = subheader_fill
            ws[f'G{grade_summary_row}'] = f'=VLOOKUP(D{grade_summary_row},ModifiedGrades!A:B,2,TRUE)'
            ws[f'G{grade_summary_row}'].font = Font(bold=True, size=12)
            ws[f'G{grade_summary_row}'].fill = subheader_fill
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        if self.modified_grade_scale:
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
        
        # Create letter grade conversion sheet
        self._add_grade_conversion_sheet(wb)
        
        # Create modified grade conversion sheet if provided
        if self.modified_grade_scale:
            self._add_modified_grade_conversion_sheet(wb)
        
        # Save workbook
        wb.save(filename)
    
    def _add_grade_conversion_sheet(self, wb: Workbook) -> None:
        """Add a sheet with letter grade conversion table.
        
        Args:
            wb: Workbook to add the sheet to
        """
        # Create new sheet
        ws = wb.create_sheet("Grades")
        
        # Get the letter grade scale
        if self.letter_grade_scale:
            scale = self.letter_grade_scale
        else:
            # Default MIT scale if none provided
            scale = {
                0.00: 'F',
                0.61: 'D',
                0.70: 'C-',
                0.74: 'C',
                0.77: 'C+',
                0.80: 'B-',
                0.84: 'B',
                0.87: 'B+',
                0.90: 'A-',
                0.94: 'A',
                0.97: 'A+'
            }
        
        # Header
        ws['A1'] = "Min Percentage"
        ws['B1'] = "Letter Grade"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # IMPORTANT: Sort by percentage ASCENDING for VLOOKUP with TRUE (approximate match)
        # VLOOKUP with TRUE requires data sorted in ascending order
        sorted_thresholds = sorted(scale.items(), reverse=False)
        
        row = 2
        for threshold, letter in sorted_thresholds:
            ws[f'A{row}'] = threshold  # Store as decimal (0.93 for 93%)
            ws[f'A{row}'].number_format = '0.00%'
            ws[f'B{row}'] = letter
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
    
    def _add_modified_grade_conversion_sheet(self, wb: Workbook) -> None:
        """Add a sheet with modified letter grade conversion table.
        
        Args:
            wb: Workbook to add the sheet to
        """
        # Create new sheet
        ws = wb.create_sheet("ModifiedGrades")
        
        # Use the modified grade scale
        scale = self.modified_grade_scale
        
        # Header
        ws['A1'] = "Min Percentage"
        ws['B1'] = "Letter Grade"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Sort by percentage ASCENDING for VLOOKUP with TRUE (approximate match)
        sorted_thresholds = sorted(scale.items(), reverse=False)
        
        row = 2
        for threshold, letter in sorted_thresholds:
            ws[f'A{row}'] = threshold  # Store as decimal (0.93 for 93%)
            ws[f'A{row}'].number_format = '0.00%'
            ws[f'B{row}'] = letter
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15

    def _get_category_weight(self, type_name: str) -> float:
        """Get the weight for a category as a percentage (0-100)."""
        # Try config first (both Canvas and manual config)
        if hasattr(self, 'config') and self.config:
            return self.config.weight_for_type(type_name) * 100
        # Fallback to assignment groups if stored
        elif hasattr(self, '_assignment_groups'):
            for group in self._assignment_groups:
                if getattr(group, 'name', None) == type_name:
                    return getattr(group, 'group_weight', 0.0)
        return 0.0
